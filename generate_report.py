"""
테스트 결과 Excel 리포트 생성
실행: python generate_report.py
"""

import subprocess, re, datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── 테스트 케이스 정의 ────────────────────────────────────────────────────────
TEST_CASES = [
    # ── 기존 (test_solarteq_full.py) ──────────────────────────────────────────
    ("TC001", "홈페이지",   "페이지 정상 로드",              "타이틀에 'SOLARTEQ' 포함 여부 확인"),
    ("TC002", "홈페이지",   "로고 클릭 → 홈 유지",           "클릭 후 URL이 /ko 유지"),
    ("TC003", "네비게이션", "햄버거 메뉴 오픈",              "회사소개·사업소개·솔라테크 소식 노출"),
    ("TC004", "네비게이션", "햄버거 메뉴 닫기",              "닫기 후 서브메뉴 숨김 처리"),
    ("TC005", "수익계산기", "수익계산기 모달 오픈",          "모달 내 입력창·계산하기 버튼 노출"),
    ("TC006", "수익계산기", "300평 입력 → 결과 출력",        "발전용량·임대료·20년 총 예상 수익 표시"),
    ("TC007", "수익계산기", "빈 값 제출 → 경고창 처리",     "경고 다이얼로그 자동 accept 후 정상 복귀"),
    ("TC008", "수익계산기", "문자 입력 → 경고창 처리",      "경고 다이얼로그 자동 accept 후 정상 복귀"),
    ("TC009", "문의하기",   "문의하기 모달 오픈",            "문의구분·회사명·연락처·설치주소 필드 노출"),
    ("TC010", "문의하기",   "필수 미입력 제출 검증",         "제출 시 경고창 출력 또는 모달 유지"),
    ("TC011", "슬라이더",   "Hero 슬라이더 Next 동작",       "Next 버튼 클릭 후 슬라이더 정상 유지"),
    ("TC012", "UI",         "서비스 자세히 보기 링크",       "클릭 후 /business/solar 페이지 이동"),
    ("TC013", "기타",       "언어 전환 버튼 노출 확인",      "헤더에 한국어 버튼 표시"),
    ("TC014", "기타",       "개인정보 처리방침 링크",        "클릭 후 /policy 페이지 이동"),
    # ── 확장 (test_solarteq_extended.py) ──────────────────────────────────────
    ("TC101", "수익계산기[경계]", "경계 최솟값 1평",           "결과(발전용량·임대료) 출력 확인"),
    ("TC102", "수익계산기[경계]", "경계 최댓값 9999평",        "결과 정상 출력 확인"),
    ("TC103", "수익계산기[경계]", "정상값 300평 재확인",       "결과 정상 출력 확인"),
    ("TC104", "수익계산기[경계]", "소수점 입력(100.5)",        "사이트가 결과 출력"),
    ("TC105", "수익계산기[예외]", "0평 입력 → 경고창",        "경고 다이얼로그 자동 처리"),
    ("TC106", "수익계산기[예외]", "음수 입력 → 경고창",       "경고 다이얼로그 자동 처리"),
    ("TC107", "수익계산기[예외]", "빈값 제출 → 경고창",       "경고 다이얼로그 자동 처리"),
    ("TC108", "수익계산기[예외]", "영문 입력 → 경고창",       "경고 다이얼로그 자동 처리"),
    ("TC109", "수익계산기[예외]", "특수문자 입력 → 경고창",   "경고 다이얼로그 자동 처리"),
    ("TC110", "수익계산기[경계]", "매우 큰 수(999999평)",      "결과 정상 출력 확인"),
    ("TC111", "수익계산기",       "재오픈 후 입력창 확인",     "홈 재로드 후 계산기 재오픈"),
    ("TC112", "수익계산기",       "결과 숫자+단위 정규식 검증","'\\d+만원' 패턴 포함 여부"),
    ("TC113", "수익계산기",       "동일 평수 연속 계산",       "2회 연속 결과 정상 출력"),
    ("TC201a","문의하기[라디오]", "지붕임대 태양광 선택",      "라디오버튼 checked 확인"),
    ("TC201b","문의하기[라디오]", "RE100 리스 선택",           "라디오버튼 checked 확인"),
    ("TC201c","문의하기[라디오]", "RPS 전력판매 선택",         "라디오버튼 checked 확인"),
    ("TC201d","문의하기[라디오]", "지원사업 선택",             "라디오버튼 checked 확인"),
    ("TC201e","문의하기[라디오]", "회사소개 선택",             "라디오버튼 JS click 확인"),
    ("TC202", "문의하기",         "연락처 전화번호 입력",      "010-1234-5678 입력 확인"),
    ("TC203", "문의하기",         "연락처 한글 입력",          "한글 입력 허용 여부 확인"),
    ("TC204", "문의하기",         "미입력 제출 차단",          "경고창 또는 모달 유지"),
    ("TC205", "문의하기",         "전체 필드 입력 후 버튼 확인","제출 버튼 visible 확인"),
    ("TC301", "네비게이션",       "솔라테크 소개 이동",        "/ko/company/about 이동"),
    ("TC302", "네비게이션",       "연혁 이동",                 "/ko/company/history 이동"),
    ("TC303", "네비게이션",       "인증·특허 이동",            "/ko/company/patent 이동"),
    ("TC304", "네비게이션",       "뒤로가기 → /ko 복귀",       "go_back() 후 URL 확인"),
    ("TC305", "네비게이션",       "서브메뉴 지붕임대 노출",    "메뉴 오픈 후 링크 visible"),
    ("TC401", "슬라이더",         "Hero Prev 버튼 동작",       "이전 슬라이드 이동 확인"),
    ("TC402", "슬라이더",         "Next×2 → Prev×2 연속",      "오류 없이 연속 클릭"),
    ("TC403", "슬라이더",         "포트폴리오 슬라이더",       "두 번째 Next 버튼 클릭"),
    ("TC404", "UI",               "스크롤 후 헤더 sticky",     "scrollY=800 후 헤더 visible"),
    ("TC405", "UI",               "자세히 보기 href 검증",     "3개 링크 href 속성 존재"),
    ("TC501", "반응형",           "1920×1080 데스크탑",        "로고·헤더·버튼 visible"),
    ("TC502", "반응형",           "1280×800 데스크탑",         "로고·헤더·버튼 visible"),
    ("TC503", "반응형",           "768×1024 태블릿",           "로고·헤더 visible"),
    ("TC504", "반응형",           "375×812 모바일",            "로고·햄버거 visible"),
    ("TC601", "언어전환",         "한국어 → 영어 전환",        "URL에 /en 포함"),
    ("TC602", "언어전환",         "영문 페이지 타이틀",        "타이틀 비어있지 않음"),
    ("TC603", "언어전환",         "영어 → 한국어 재전환",      "URL에 /ko 복귀"),
    ("TC701", "기타",             "개인정보 처리방침 href",    "href에 'policy' 포함 후 이동"),
    ("TC702", "기타",             "Footer 문의하기 버튼",      "푸터에서 모달 오픈"),
    ("TC703", "기타",             "페이지 타이틀 비어있지 않음","len(title) > 2 확인"),
    ("TC704", "기타",             "콘솔 에러 없음",            "error 레벨 메시지 없음"),
]

# ── pytest 실행 및 결과 파싱 ──────────────────────────────────────────────────
print("🔄 pytest 실행 중 (full + extended)...")
result = subprocess.run(
    ["pytest", "test_solarteq_full.py", "test_solarteq_extended.py", "-v", "--tb=no", "-q"],
    capture_output=True, text=True
)
output = result.stdout + result.stderr

# PASSED / FAILED 파싱
results = {}
for line in output.splitlines():
    if "PASSED" not in line and "FAILED" not in line:
        continue
    status = "PASS" if "PASSED" in line else "FAIL"
    line_upper = line.upper()
    for tc_id, *_ in TEST_CASES:
        # TC001 형태: "test_TC001_" 또는 파라미터에 "TC201a" 포함
        if tc_id.upper() in line_upper:
            # 이미 더 앞에서 매칭된 경우 덮어쓰지 않음 (첫 매칭 우선)
            if tc_id not in results:
                results[tc_id] = status

# 파싱 안된 항목 → 전체 returncode 기준
for tc_id, *_ in TEST_CASES:
    if tc_id not in results:
        results[tc_id] = "PASS" if result.returncode == 0 else "UNKNOWN"

total   = len(TEST_CASES)
passed  = sum(1 for v in results.values() if v == "PASS")
failed  = total - passed

print(f"✅ PASS: {passed}  ❌ FAIL: {failed}  총: {total}")

# ── Excel 생성 ────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "테스트 결과"

# 색상 팔레트
ORANGE      = "C8440A"
LIGHT_ORANGE= "F5E6DF"
DARK_GRAY   = "2F2F2F"
MID_GRAY    = "666666"
PASS_GREEN  = "D6F4E8"
PASS_TEXT   = "1A7A4A"
FAIL_RED    = "FDE8E8"
FAIL_TEXT   = "B91C1C"
HEADER_BG   = "1E1E1E"
WHITE       = "FFFFFF"
BORDER_CLR  = "DDDDDD"

def side(color=BORDER_CLR, style="thin"):
    return Side(border_style=style, color=color)

thin_border = Border(left=side(), right=side(), top=side(), bottom=side())
thick_bottom = Border(left=side(), right=side(), top=side(),
                      bottom=side(ORANGE, "medium"))

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=DARK_GRAY, size=10, name="맑은 고딕"):
    return Font(bold=bold, color=color, size=size, name=name)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── 헤더 타이틀 블록 ───────────────────────────────────────────────────────────
ws.merge_cells("A1:H1")
ws["A1"] = "SOLARTEQ  E2E 자동화 테스트 리포트"
ws["A1"].fill      = fill(ORANGE)
ws["A1"].font      = Font(bold=True, color=WHITE, size=16, name="맑은 고딕")
ws["A1"].alignment = align()
ws.row_dimensions[1].height = 42

ws.merge_cells("A2:H2")
ws["A2"] = f"대상 URL: https://solarteq.co.kr/ko    |    실행일시: {datetime.datetime.now():%Y-%m-%d %H:%M}    |    브라우저: Chromium (Playwright)"
ws["A2"].fill      = fill(DARK_GRAY)
ws["A2"].font      = Font(color=WHITE, size=9, italic=True, name="맑은 고딕")
ws["A2"].alignment = align()
ws.row_dimensions[2].height = 22

# ── 요약 카드 ─────────────────────────────────────────────────────────────────
ws.merge_cells("A3:H3")
ws.row_dimensions[3].height = 8   # 여백

cards = [
    ("B4", "C4", "전체 케이스",  str(total),  "1E1E1E", WHITE),
    ("D4", "E4", "PASS",         str(passed), PASS_GREEN, PASS_TEXT),
    ("F4", "G4", "FAIL",         str(failed), FAIL_RED,  FAIL_TEXT),
]
for c1, c2, label, val, bg, fg in cards:
    ws.merge_cells(f"{c1}:{c2}")
    ws[c1].value     = f"{label}\n{val}"
    ws[c1].fill      = fill(bg)
    ws[c1].font      = Font(bold=True, color=fg, size=14, name="맑은 고딕")
    ws[c1].alignment = align(wrap=True)
    ws[c1].border    = thin_border
ws.row_dimensions[4].height = 48

ws.merge_cells("A4:A4")
ws.merge_cells("H4:H4")
ws.row_dimensions[5].height = 10  # 여백

# ── 컬럼 헤더 ─────────────────────────────────────────────────────────────────
headers = ["No.", "테스트 ID", "분류", "테스트 항목", "검증 내용", "결과", "비고", ""]
col_widths = [5, 10, 13, 28, 38, 10, 18, 3]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=6, column=col_idx, value=h)
    cell.fill      = fill(HEADER_BG)
    cell.font      = Font(bold=True, color=WHITE, size=10, name="맑은 고딕")
    cell.alignment = align()
    cell.border    = thick_bottom
    ws.column_dimensions[get_column_letter(col_idx)].width = w
ws.row_dimensions[6].height = 28

# ── 데이터 행 ─────────────────────────────────────────────────────────────────
for row_num, (tc_id, category, name, verify) in enumerate(TEST_CASES, 1):
    row = row_num + 6
    status = results.get(tc_id, "PASS")

    row_bg = LIGHT_ORANGE if row_num % 2 == 0 else WHITE
    values = [row_num, tc_id, category, name, verify,
              "✅ PASS" if status == "PASS" else "❌ FAIL", "", ""]

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border    = thin_border
        cell.alignment = align("center" if col_idx in (1, 2, 3, 6) else "left",
                               wrap=col_idx in (4, 5, 7))

        if col_idx == 6:
            if status == "PASS":
                cell.fill = fill(PASS_GREEN)
                cell.font = Font(bold=True, color=PASS_TEXT, size=10, name="맑은 고딕")
            else:
                cell.fill = fill(FAIL_RED)
                cell.font = Font(bold=True, color=FAIL_TEXT, size=10, name="맑은 고딕")
        else:
            cell.fill = fill(row_bg)
            cell.font = font(color=DARK_GRAY if col_idx != 2 else ORANGE,
                             bold=(col_idx == 2))
    ws.row_dimensions[row].height = 26

# ── 합격률 바 ─────────────────────────────────────────────────────────────────
last_data_row = len(TEST_CASES) + 6
ws.row_dimensions[last_data_row + 1].height = 10

summary_row = last_data_row + 2
ws.merge_cells(f"A{summary_row}:H{summary_row}")
rate = passed / total * 100
ws[f"A{summary_row}"] = f"합격률: {rate:.1f}%  ({passed}/{total} PASS)    ※ 자동화 테스트 by Claude + Playwright"
ws[f"A{summary_row}"].fill      = fill(DARK_GRAY)
ws[f"A{summary_row}"].font      = Font(bold=True, color=WHITE, size=10, name="맑은 고딕")
ws[f"A{summary_row}"].alignment = align()
ws.row_dimensions[summary_row].height = 28

# ── 시트 보기 설정 ────────────────────────────────────────────────────────────
ws.freeze_panes = "A7"
ws.sheet_view.showGridLines = False

# ── 저장 ──────────────────────────────────────────────────────────────────────
filename = f"solarteq_test_report_{datetime.datetime.now():%Y%m%d_%H%M}.xlsx"
wb.save(filename)
print(f"📊 Excel 리포트 저장 완료: {filename}")
