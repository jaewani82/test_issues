"""
솔라테크 E2E 테스트 결과 Excel 리포트 생성
TC 번호: TC001 ~ TC057 순차 부여
실행: python make_report_now.py
"""

import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 테스트 데이터 (카테고리, 항목, 검증내용, 결과) ────────────────────────────
# 섹션 헤더는 cat이 "──" 로 시작하는 행으로 구분
RAW = [
    # section
    ("§", "홈페이지", "", ""),
    ("홈페이지",       "페이지 정상 로드",               "타이틀에 'SOLARTEQ' 포함 여부",           "PASS"),
    ("홈페이지",       "로고 클릭 → 홈 유지",            "클릭 후 URL /ko 유지",                    "PASS"),
    # section
    ("§", "네비게이션", "", ""),
    ("네비게이션",     "햄버거 메뉴 오픈",               "회사소개·사업소개·소식 노출",              "PASS"),
    ("네비게이션",     "햄버거 메뉴 닫기",               "닫기 후 서브메뉴 숨김",                    "PASS"),
    ("네비게이션",     "솔라테크 소개 이동",             "/ko/company/about 이동",                  "PASS"),
    ("네비게이션",     "연혁 이동",                     "/ko/company/history 이동",                 "PASS"),
    ("네비게이션",     "인증·특허 이동",                "/ko/company/patent 이동",                  "PASS"),
    ("네비게이션",     "뒤로가기 → /ko 복귀",           "go_back() 후 URL 확인",                    "PASS"),
    ("네비게이션",     "서브메뉴 지붕임대 노출",         "메뉴 오픈 후 링크 visible",                "PASS"),
    # section
    ("§", "수익계산기 — 기본", "", ""),
    ("수익계산기",     "수익계산기 모달 오픈",           "입력창·계산하기 버튼 노출",                "PASS"),
    ("수익계산기",     "300평 입력 → 결과 출력",         "발전용량·임대료·총수익 표시",              "PASS"),
    ("수익계산기",     "빈 값 제출 → 경고창",            "다이얼로그 자동 accept",                   "PASS"),
    ("수익계산기",     "문자 입력 → 경고창",             "다이얼로그 자동 accept",                   "PASS"),
    ("수익계산기",     "재오픈 후 입력창 확인",          "홈 재로드 후 재오픈",                      "PASS"),
    ("수익계산기",     "결과 숫자+단위 정규식 검증",     r"\d+만원 패턴 존재 확인",                  "PASS"),
    ("수익계산기",     "동일 평수 연속 계산",            "2회 연속 결과 정상 출력",                  "PASS"),
    # section
    ("§", "수익계산기 — 경계값/예외", "", ""),
    ("수익계산기[경계]","경계 최솟값 1평",               "결과 정상 출력",                           "PASS"),
    ("수익계산기[경계]","경계 최댓값 9999평",            "결과 정상 출력",                           "PASS"),
    ("수익계산기[경계]","정상값 300평 재확인",           "결과 정상 출력",                           "PASS"),
    ("수익계산기[경계]","소수점 입력(100.5평)",          "사이트가 결과 출력",                       "PASS"),
    ("수익계산기[경계]","매우 큰 수(999999평)",          "결과 정상 출력",                           "PASS"),
    ("수익계산기[예외]","0평 입력 → 경고창",            "다이얼로그 자동 처리",                     "PASS"),
    ("수익계산기[예외]","음수 입력 → 경고창",           "다이얼로그 자동 처리",                     "PASS"),
    ("수익계산기[예외]","빈값 제출 → 경고창",           "다이얼로그 자동 처리",                     "PASS"),
    ("수익계산기[예외]","영문 입력 → 경고창",           "다이얼로그 자동 처리",                     "PASS"),
    ("수익계산기[예외]","특수문자 입력 → 경고창",       "다이얼로그 자동 처리",                     "PASS"),
    # section
    ("§", "문의하기 모달", "", ""),
    ("문의하기",       "문의하기 모달 오픈",             "4개 필드 노출 확인",                       "PASS"),
    ("문의하기",       "필수 미입력 제출 검증",          "경고창 출력 또는 모달 유지",               "PASS"),
    ("문의하기[라디오]","지붕임대 태양광 선택",          "radio checked 확인",                       "PASS"),
    ("문의하기[라디오]","RE100 리스 선택",               "radio checked 확인",                       "PASS"),
    ("문의하기[라디오]","RPS 전력판매 선택",             "radio checked 확인",                       "PASS"),
    ("문의하기[라디오]","지원사업 선택",                 "radio checked 확인",                       "PASS"),
    ("문의하기[라디오]","회사소개·지명원 요청 선택",     "JS click 우회",                            "PASS"),
    ("문의하기",       "연락처 전화번호 입력",           "010-1234-5678 입력",                       "PASS"),
    ("문의하기",       "연락처 한글 입력",               "입력 허용 여부 확인",                      "PASS"),
    ("문의하기",       "미입력 제출 차단",               "경고창 또는 모달 유지",                    "PASS"),
    ("문의하기",       "전체 필드 입력 후 버튼 확인",    "제출 버튼 visible 확인",                   "PASS"),
    # section
    ("§", "슬라이더 / UI", "", ""),
    ("슬라이더",       "Hero 슬라이더 Next",             "Next 클릭 후 슬라이더 유지",               "PASS"),
    ("슬라이더",       "Hero Prev 버튼 동작",            "이전 슬라이드 이동",                       "PASS"),
    ("슬라이더",       "Next×2 → Prev×2 연속",          "오류 없이 연속 클릭",                      "PASS"),
    ("슬라이더",       "포트폴리오 슬라이더 Next",       "두 번째 Next 버튼 클릭",                   "PASS"),
    ("UI",             "자세히 보기 링크",               "/business/solar 이동",                    "PASS"),
    ("UI",             "스크롤 후 헤더 sticky",          "scrollY=800 후 헤더 visible",              "PASS"),
    ("UI",             "자세히 보기 href 검증",          "3개 링크 href 속성 존재",                  "PASS"),
    # section
    ("§", "반응형 Viewport", "", ""),
    ("반응형",         "1920×1080 데스크탑",             "로고·헤더·버튼 visible",                   "PASS"),
    ("반응형",         "1280×800 데스크탑",              "로고·헤더·버튼 visible",                   "PASS"),
    ("반응형",         "768×1024 태블릿",                "로고·헤더 visible",                        "PASS"),
    ("반응형",         "375×812 모바일",                 "로고·햄버거 visible",                      "PASS"),
    # section
    ("§", "언어 전환", "", ""),
    ("언어전환",       "한국어 → 영어 전환",             "URL에 /en 포함",                           "PASS"),
    ("언어전환",       "영문 페이지 타이틀",             "타이틀 비어있지 않음",                     "PASS"),
    ("언어전환",       "영어 → 한국어 재전환",           "URL에 /ko 복귀",                           "PASS"),
    # section
    ("§", "기타 / 보완", "", ""),
    ("기타",           "언어 전환 버튼 노출",            "헤더 한국어 버튼 표시",                    "PASS"),
    ("기타",           "개인정보 처리방침 링크",         "/ko/policy 이동",                         "PASS"),
    ("기타",           "개인정보 처리방침 href",         "href에 'policy' 포함 후 이동",             "PASS"),
    ("기타",           "Footer 문의하기 버튼",           "푸터에서 모달 오픈",                       "PASS"),
    ("기타",           "페이지 타이틀 비어있지 않음",    "len(title) > 2",                           "PASS"),
    ("기타",           "콘솔 에러 없음",                 "error 레벨 메시지 없음",                   "PASS"),
]

# 섹션 행 제외 후 순차 번호 부여
TC_ROWS  = [(cat, name, verify, status) for cat, name, verify, status in RAW if cat != "§"]
total    = len(TC_ROWS)
passed   = sum(1 for *_, s in TC_ROWS if s == "PASS")
failed   = total - passed

# ── 스타일 상수 ───────────────────────────────────────────────────────────────
ORANGE   = "C8440A"
DARK     = "1E1E1E"
WHITE    = "FFFFFF"
PASS_BG  = "D6F4E8"
PASS_FG  = "1A7A4A"
FAIL_BG  = "FDE8E8"
FAIL_FG  = "B91C1C"
SEC_BG   = "F0E8E2"
SEC_FG   = "8B3A0A"
ROW_A    = "FAFAFA"
ROW_B    = "F3EBE6"
BORDER_C = "DDDDDD"

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, color=DARK, size=10):
    return Font(bold=bold, color=color, size=size, name="맑은 고딕")
def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr():
    s = Side(border_style="thin", color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)
def hdr_bdr():
    s = Side(border_style="thin",   color=BORDER_C)
    b = Side(border_style="medium", color=ORANGE)
    return Border(left=s, right=s, top=s, bottom=b)

# ── Workbook 생성 ─────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "테스트 결과"
ws.sheet_view.showGridLines = False

# 컬럼 너비: No. | TC ID | 분류 | 테스트 항목 | 검증 내용 | 결과 | 여백
for i, w in enumerate([5, 10, 20, 30, 38, 10, 3], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── 타이틀 ────────────────────────────────────────────────────────────────────
ws.merge_cells("A1:G1")
ws["A1"] = "SOLARTEQ  E2E 자동화 테스트 리포트"
ws["A1"].fill      = fill(ORANGE)
ws["A1"].font      = Font(bold=True, color=WHITE, size=17, name="맑은 고딕")
ws["A1"].alignment = aln()
ws.row_dimensions[1].height = 44

ws.merge_cells("A2:G2")
ws["A2"] = (f"대상 URL: https://solarteq.co.kr/ko    |    "
            f"실행일시: {datetime.datetime.now():%Y-%m-%d %H:%M}    |    "
            f"브라우저: Chromium (Playwright)")
ws["A2"].fill      = fill(DARK)
ws["A2"].font      = Font(color=WHITE, size=9, italic=True, name="맑은 고딕")
ws["A2"].alignment = aln()
ws.row_dimensions[2].height = 22

# ── 요약 카드 ─────────────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 8
for (c1, c2), label, val, bg, fg in [
    (("B4","C4"), "전체 TC",  str(total),  DARK,    WHITE),
    (("D4","E4"), "✅  PASS", str(passed), PASS_BG, PASS_FG),
    (("F4","F4"), "❌  FAIL", str(failed), FAIL_BG, FAIL_FG),
]:
    ws.merge_cells(f"{c1}:{c2}")
    ws[c1].value     = f"{label}\n{val}"
    ws[c1].fill      = fill(bg)
    ws[c1].font      = Font(bold=True, color=fg, size=15, name="맑은 고딕")
    ws[c1].alignment = aln(wrap=True)
    ws[c1].border    = bdr()
ws.row_dimensions[4].height = 52
ws.row_dimensions[5].height = 10

# ── 합격률 바 ─────────────────────────────────────────────────────────────────
ws.merge_cells("A6:G6")
rate = passed / total * 100
ws["A6"] = (f"합격률  {rate:.1f}%   ({passed} / {total} PASS)"
            f"   ─   자동화 테스트: Claude Code + Playwright + pytest")
ws["A6"].fill      = fill(DARK)
ws["A6"].font      = Font(bold=True, color=WHITE, size=10, name="맑은 고딕")
ws["A6"].alignment = aln()
ws.row_dimensions[6].height = 26
ws.row_dimensions[7].height = 6

# ── 컬럼 헤더 ─────────────────────────────────────────────────────────────────
for ci, h in enumerate(["No.", "TC ID", "분류", "테스트 항목", "검증 내용", "결과", ""], 1):
    c = ws.cell(row=8, column=ci, value=h)
    c.fill      = fill(DARK)
    c.font      = Font(bold=True, color=WHITE, size=10, name="맑은 고딕")
    c.alignment = aln()
    c.border    = hdr_bdr()
ws.row_dimensions[8].height = 28

# ── 데이터 + 섹션 행 ─────────────────────────────────────────────────────────
excel_row = 9
tc_seq    = 0       # TC001 ~ 순차

for cat, name, verify, status in RAW:

    # 섹션 구분 행
    if cat == "§":
        ws.merge_cells(f"A{excel_row}:G{excel_row}")
        ws[f"A{excel_row}"].value     = f"   {name}"
        ws[f"A{excel_row}"].fill      = fill(SEC_BG)
        ws[f"A{excel_row}"].font      = Font(bold=True, color=SEC_FG, size=9, name="맑은 고딕")
        ws[f"A{excel_row}"].alignment = aln("left")
        ws[f"A{excel_row}"].border    = bdr()
        ws.row_dimensions[excel_row].height = 19
        excel_row += 1
        continue

    # TC 데이터 행
    tc_seq += 1
    tc_id  = f"TC{tc_seq:03d}"
    bg     = ROW_A if tc_seq % 2 == 0 else ROW_B

    vals = [tc_seq, tc_id, cat, name, verify,
            "✅  PASS" if status == "PASS" else "❌  FAIL", ""]

    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=excel_row, column=ci, value=v)
        c.border    = bdr()
        c.alignment = aln("center" if ci in (1, 2, 6) else "left",
                          wrap=ci in (3, 4, 5))
        if ci == 6:
            c.fill = fill(PASS_BG if status == "PASS" else FAIL_BG)
            c.font = Font(bold=True,
                          color=PASS_FG if status == "PASS" else FAIL_FG,
                          size=10, name="맑은 고딕")
        elif ci == 2:
            c.fill = fill(bg)
            c.font = Font(bold=True, color=ORANGE, size=9, name="맑은 고딕")
        else:
            c.fill = fill(bg)
            c.font = fnt()

    ws.row_dimensions[excel_row].height = 24
    excel_row += 1

# ── 푸터 ──────────────────────────────────────────────────────────────────────
ws.row_dimensions[excel_row].height = 8
excel_row += 1
ws.merge_cells(f"A{excel_row}:G{excel_row}")
ws[f"A{excel_row}"] = ("파일:  test_solarteq_full.py (기본 14 TC)  |  "
                        "test_solarteq_extended.py (확장 43 TC)  |  "
                        "HANDOFF.md")
ws[f"A{excel_row}"].fill      = fill(DARK)
ws[f"A{excel_row}"].font      = Font(color=WHITE, size=8, italic=True, name="맑은 고딕")
ws[f"A{excel_row}"].alignment = aln()
ws.row_dimensions[excel_row].height = 22

ws.freeze_panes = "A9"

# ── 저장 ──────────────────────────────────────────────────────────────────────
fname = f"solarteq_report_final_{datetime.datetime.now():%Y%m%d_%H%M}.xlsx"
wb.save(fname)
print(f"\n📊 리포트 생성 완료: {fname}")
print(f"   총 {total}개 TC  |  PASS {passed}  |  FAIL {failed}  |  합격률 {rate:.1f}%")
print(f"   번호 범위: TC001 ~ TC{total:03d}")
